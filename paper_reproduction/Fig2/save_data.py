#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 10 01:17:34 2023

@author: nel
"""

import pandas as pd
import openpyxl
import glob
import os
import numpy as np

#%% save function
def multiple_dfs(df_list, sheets, file_name, spaces, text):
    try:
        book=openpyxl.load_workbook(file_name)
        print("existing workbook")
    except:
        book=openpyxl.Workbook()
        print("new workbook")
        book.save(file_name)
        
    writer = pd.ExcelWriter(file_name, engine="openpyxl")
    writer.book = book
    writer.sheets = {ws.title: ws for ws in sheets}
    print(writer.sheets)
    
    for i,dataframe in enumerate(df_list):
        row=3
        dataframe.to_excel(writer,sheet_name=sheets[i], startrow=row, startcol=0, index=False, na_rep="NA")
        row = row+ len(dataframe.index)+ spaces+ 1
        # writer.sheets[sheets[i]].append(list(dataframe["times"][0]))
        writer.sheets[sheets[i]].cell(1,1).value=text
        writer.save()
#%% 3b
# lh_nnls_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/DATA_PAPER_ELIFE/*/nnls.npy")
lh_nnls_files = sorted(glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/data/voltage_data/*/nnls.npy"))
# nnls_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/DATA_PAPER_ELIFE/*/v_nnls_*.npy")
nnls_files = sorted(glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/data/voltage_data/*/v_nnls_*.npy"))
#%% uncomment this cell to add all traces (not recommended)
# dfb = pd.DataFrame()
# max_len = 20000
# for f in sorted(lh_nnls_files):
#     dat = np.load(f, allow_pickle=True)
#     print(dat.shape)
#     for i in range(dat.shape[0]):
#         cell_dat = dat[i]
#         filler = [None]*(max_len-len(cell_dat))
#         dfb[f.split("/")[-2]+str(i)+ "_lawson_hanson"] = np.concatenate((cell_dat, filler))
# for f in sorted(nnls_files):
#     dat = np.load(f, allow_pickle=True)
#     print(dat.shape)
#     for i in range(dat.shape[0]):
#         cell_dat = dat[i]
#         filler = [None]*(max_len-len(cell_dat))
#         dfb[f.split("/")[-2]+ "_"+ f.split("_")[-1][:-4] +"iters"] = np.concatenate((cell_dat, filler))
#%% 3b, voltage
from scipy.stats import pearsonr
dfb = pd.DataFrame()
max_len = 55
for i, lf in enumerate(lh_nnls_files):
    lh_dat = np.load(lf, allow_pickle=True)
    for ff in nnls_files[i*3:i*3+3]:
        fi_dat = np.load(ff, allow_pickle=True)
        corrs = []
        print(fi_dat.shape, lh_dat.shape)
        for j in range(lh_dat.shape[0]):
            corrs.append(pearsonr(fi_dat[j], lh_dat[j])[0])
        filler = [None] * (max_len - fi_dat.shape[0])
        dfb[ff.split("voltage_data/")[-1]] = np.concatenate((corrs, filler))
        
#%% 3b, c, e trace data
snr_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_SNR.npy")
rscore_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_rscore.npy")
incl_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_incl.npy")
dfc = pd.DataFrame()
max_len = 485
for f in sorted(snr_files):
    dat = np.load(f, allow_pickle=True)
    print(dat.shape)
    filler = [None]*(max_len-dat.shape[0])
    dfc[f.split("/")[-1][:-4]] = np.concatenate((dat, filler))
for f in sorted(rscore_files)[:-1]:
    dat = np.load(f, allow_pickle=True)
    print(dat.shape)
    filler = [None]*(max_len-dat.shape[0])
    dfc[f.split("/")[-1][:-4]] = np.concatenate((dat, filler))
for f in sorted(incl_files)[:-1]:
    dat = np.load(f,  allow_pickle=True)
    print(dat.shape)
    filler = [None]*(max_len-dat.shape[0])
    dfc[f.split("/")[-1][:-4]] = np.concatenate((dat, filler))
#%% 3e iteration timings
base_file  = "/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/"
dfe = pd.DataFrame()
for i,f in enumerate(["N00", "N01","N02","N03","N04","YST"]):  
    for t in ["5", "10", "30"]:
        ti = np.load(base_file + f+ "_nnls_" + t + "_time.npy")
        dfe[" ".join([f, t, "iters timing"])] = ti
#%% 3d
k53_nnls_time_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_*_*_nnls_time.npy")
cm_nnls_time_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_cm_*_nnls_time.npy")
dfd = pd.DataFrame()
max_len = 2999
for f in sorted(k53_nnls_time_files):
    dat = np.load(f, allow_pickle=True)
    print(dat.shape)
    filler = [None]*(max_len-dat.shape[0])
    dfd[f.split("/")[-1][:-4]] = np.concatenate((dat, filler))
for f in sorted(cm_nnls_time_files):
    dat = np.load(f, allow_pickle=True)[()]["T_track"][3000:-1]
    print(dat.shape)
    filler = [None]*(max_len-dat.shape[0])
    dfd[f.split("/")[-1][:-4]] = np.concatenate((dat, filler))
#%% files for 3e
iter_nnls_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_nnls_*_time.npy")
fi_nnls_files = glob.glob('/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/DATA_PAPER_ELIFE/*/v_nnls_*.npy')
#%% 3e processing omitted = > all files  used  for 3d  were  saved in  previous  sheets
#%% Fig  3  save  all
excel_folder = "../../../../media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Data"
text = "NNLS benchmarking"
sheets = ["3b"]
dfs = [dfe]
excel_name = os.path.join(excel_folder, "FIOLA3.1.xlsx")
multiple_dfs(dfs, sheets, excel_name, 2, text)

#%% files for 5a
fiola_all = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/*/*_times.npy")
fiola_crop = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/Crop/time*.npy")
cm_times = sorted(glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/CMTimes/7apr/cm_*.npy"))
fiola_batch = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/Batch/Neurs/times_*.npy")
def filter_helper(path):
    dims = ["256","768", "crop", "_C"]
    return all(x not in path for x in dims)
all_files = []
all_files += sorted(list(filter(lambda x: filter_helper(x), fiola_all)), key=lambda y: (int(y.split("/")[-2]), int(y.split("_")[-2])))
all_files += sorted(list(filter(lambda x: filter_helper(x), fiola_crop)), key=lambda y: (int(y.split("_")[-1][:-4]), int(y.split("_")[-2])))
all_files += sorted(list(filter(lambda x: filter_helper(x), fiola_batch)), key=lambda y: (int(y.split("_")[-2]), int(y.split("_")[-1][:-4])))
all_files += sorted(list(filter(lambda x: filter_helper(x), cm_times)))
all_files_headers = [f.split("/")[-1] for f in all_files]
values = []
for fl in all_files:
    if "cm" in fl:
        v = np.load(fl, allow_pickle=True)[()]
        values.append(v["T_motion"]+ v["T_track"]+ v["T_shapes"])
    else:
        values.append(np.load(fl, allow_pickle=True))
#%%
max_len = 2999
dfa = pd.DataFrame()
dfb = pd.DataFrame()
for i,val in enumerate(all_files_headers):
    if len(values[i]) > max_len:
        values[i] = values[i][3000:]
    filler = [None]* (max_len-len(values[i]))
    dfa[val] = np.concatenate((values[i], filler))
    
   
offset = 6 
for i,val in enumerate(all_files_headers[offset:offset*2]):
    j = i + offset
    if len(values[j]) > max_len:
        values[j] = values[j][3000:]
    filler = [None]* (max_len-len(values[j]))
    dfb[val] = np.concatenate((values[j], filler))

dfs = [dfa, dfb]

#%% files for 5c
all_files = ['/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_512_mc.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_1024_mc.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_512_100_nnls_time.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_512_500_nnls_time.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_1024_100_nnls_time.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_1024_500_nnls_time.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/512_100_deconv.npy',
                '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/512_500_deconv.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/1024_100_deconv.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/1024_500_deconv.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_512_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_512_500.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_1024_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_1024_500.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_512_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_512_500.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_1024_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_1024_500.npy']
all_files_headers = [fl.split("/")[-1] for fl in all_files]
values = [np.load(fl, allow_pickle=True) for fl in all_files]
#%%
max_len = 3000
dfc = pd.DataFrame()
for i,val in enumerate(all_files_headers):

    filler = [None]* (max_len-len(values[i]))
    dfc[val] = np.concatenate((values[i], filler))
dfs.append(dfc)
  
#%%
excel_folder = "../../../../media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Data"
text = "Timing data for fiola"
sheets = ["5a", "5b", "5c"]
excel_name = os.path.join(excel_folder, "FIOLA5_1.xlsx")
multiple_dfs(dfs, sheets, excel_name, 2, "test")

#%% fig 2c
df2c = pd.DataFrame()
filler = [None]*3000
stats = {}
base_file = '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_'
nmes = ["512_cm", "512_fi", "512_fc", "512_s2", "512_s2r",
        "1024_cm", "1024_fi", "1024_fc", "1024_s2", "1024_s2r"]
for n in nmes:
    if "s" in n:
        temp="test_19-09-22/"
        vals=1000*np.array(np.load(base_file + temp + n +
                                            ".npy", allow_pickle=True))[0:]
        
    else:
        if "cm" in n:
            multiplier = 1
            offset=1
        else:
            multiplier = 1000
            offset=1500
        vals = multiplier * \
            np.array(np.load(base_file + n + ".npy",
                     allow_pickle=True)[()]["mc"][1:])
    df2c[n] = np.concatenate([vals, filler[len(vals):]])
    print(n, df2c[n].size)
#%%  write 2c to excel
excel_folder = "../../../../media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Data"
text = "Timing data for fiola motion correction"
sheets = ["2c"]
excel_name = os.path.join(excel_folder, "FIOLA2.xlsx")
df2c.columns = ["512x512_caiman","512x512_fiola", "512x512_croppedFiola", "512x512_s2pNonrigid", "512x512_s2pRigid", "1024x1024_caiman","1024x1024_fiola", "1024x1024_croppedFiola", "1024x1024_s2pNonrigid", "1024x1024_s2pRigid"]
multiple_dfs([df2c], sheets, excel_name, 2, text)

