#!/usr/bin/env python

#%% save the result to excel file
import pandas as pd
import openpyxl
def multiple_dfs(df_list, sheets, file_name, spaces, text, row = 2):
    try:
        book = openpyxl.load_workbook(file_name)
        print('existing workbook')
    except:
        book = openpyxl.Workbook()
        print('new workbook')

    writer = pd.ExcelWriter(file_name,engine='openpyxl') 
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=row, startcol=0, index=False, na_rep='NA')   
        row = row + len(dataframe.index) + spaces + 1
    #writer.sheets[sheets].cell(1,1).style.alignment.wrap_text = True
    if text is not None:
        writer.sheets[sheets].cell(1,1).value = text
    writer.save()
    
# excel_folder = '/home/nel/NEL-LAB Dropbox/NEL/Papers/VolPy/Figures/plos/excel_data'
# df = pd.DataFrame({'file':file_names,'f1 score':values})
# dfs = [df]
# text = 'F1 score of VolPy for all evaluated datasets in validation.'
# fig_name = 'Fig 3b'
# excel_name = os.path.join(excel_folder, 'volpy_data.xlsx')
# # run function
# multiple_dfs(dfs, fig_name, excel_name, 2, text)

#%%
# labels = ['VolPy', 'CaImAn', 'SGPMD-NMF', 'Suite2p', 'PCA-ICA', 'MeanROI']
# results = result_all.copy()
# df1 = pd.DataFrame({})
# for idx1, s1 in enumerate(labels):
#     for idx2, s2 in enumerate(x):
#         rr = results[idx1].item()['result'][idx2]['F1'].copy()
#         if len(rr) < 10:
#             rr = rr+ [np.nan] * (10-len(rr))
#         df1[s1 + '_' + str(s2)] = rr

# df2 = pd.DataFrame({'spike amplitude':x,'VolPy':data_points[0], 'CaImAn':data_points[1], 
#                    'SGPMD-NMF':data_points[2], 'Suite2p':data_points[3], 
#                    'PCA-ICA':data_points[4], 'MeanROI':data_points[5]})

# dfs = [df1,df2]
# text = 'Average F1 score against ground truth in function of spike amplitude. All algorithms (including VolPy) were evaluated with the optimal threshold.'
# fig_name = 'Fig 4c left'
# excel_name = os.path.join(excel_folder, 'volpy_data.xlsx')# run function
# multiple_dfs(dfs, fig_name, excel_name, 2, text)