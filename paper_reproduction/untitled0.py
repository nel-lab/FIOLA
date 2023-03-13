#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 10 01:17:34 2023

@author: nel
"""

import pandas as pd
import openpyxl
import glob
import os
import numpy as np
#%% files for 3
#%% 3b
lh_nnls_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/DATA_PAPER_ELIFE/*/nnls.npy")
lh_nnls_files += glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/data/voltage_data/*/nnls.npy")
nnls_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/DATA_PAPER_ELIFE/*/v_nnls_*.npy")
nnls_files += glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/data/voltage_data/*/v_nnls_*.npy")
#%%
dfb = pd.DataFrame()
max_len = 20000
for f in sorted(lh_nnls_files):
    dat = np.load(f, allow_pickle=True)
    print(dat.shape)
    for i in range(dat.shape[0]):
        cell_dat = dat[i]
        filler = [None]*(max_len-len(cell_dat))
        dfb[f.split("/")[-2]+str(i)+ "_lawson_hanson"] = np.concatenate((cell_dat, filler))
for f in sorted(nnls_files):
    dat = np.load(f, allow_pickle=True)
    print(dat.shape)
    for i in range(dat.shape[0]):
        cell_dat = dat[i]
        filler = [None]*(max_len-len(cell_dat))
        dfb[f.split("/")[-2]+ "_"+ f.split("_")[-1][:-4] +"iters"] = np.concatenate((cell_dat, filler))
    
#%% 3c
snr_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_SNR.npy")
rscore_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_rscore.npy")

#%% 3d
k53_nnls_time_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_*_nnls_time.npy")
cm_nnls_time_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_*_cm.npy")
#%% 3e
time_nnls_files = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/CalciumComparison/*_nnls_*_time.npy")
nnls_files += glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/data/voltage_data/*/v_nnls_*.npy")

#%% files for 5a
fiola_all = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/*/*_times.npy")
fiola_crop = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/Crop/time*.npy")
cm_times = sorted(glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/CMTimes/26feb_finalsub/cm_*.npy"))
fiola_batch = glob.glob("/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/FastResults/Batch/Neurs/times_*.npy")
def filter_helper(dims, path):
    return all(x not in path for x in dims)
all_files = []
all_files += sorted(list(filter(lambda x: filter_helper(["256","768", "crop"], x), fiola_all)), key=lambda y: (int(y.split("/")[-2]), int(y.split("_")[-2])))
all_files += sorted(list(filter(lambda x: filter_helper(["256","768"], x), fiola_crop)), key=lambda y: (int(y.split("_")[-1][:-4]), int(y.split("_")[-2])))
all_files += sorted(list(filter(lambda x: filter_helper(["256","768"], x), fiola_batch)), key=lambda y: (int(y.split("_")[-2]), int(y.split("_")[-1][:-4])))
all_files += sorted(list(filter(lambda x: filter_helper(["256","768"], x), cm_times)))
all_files_headers = [f.split("/")[-1] for f in all_files]
values = [np.load(fl, allow_pickle=True) for fl in all_files]
#%%
max_len = 2997
dfa = pd.DataFrame()
dfb = pd.DataFrame()
for i,val in enumerate(all_files_headers):
    if len(values[i]) > max_len:
        values[i] = values[i][3000:]
    filler = [None]* (max_len-len(values[i]))
    dfa[val] = np.concatenate((values[i], filler))
    
   
offset = 6 
for i,val in enumerate(all_files_headers[offset:offset*2]):
    j = i + offset
    if len(values[j]) > max_len:
        values[j] = values[j][3000:]
    filler = [None]* (max_len-len(values[j]))
    dfb[val] = np.concatenate((values[j], filler))

dfs = [dfa, dfb]

#%% files for 5c
all_files = ['/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_512_mc.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_1024_mc.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_512_100_nnls_time.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_512_500_nnls_time.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_1024_100_nnls_time.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/MotCorr/suite2p_shifts/k53_fi_1024_500_nnls_time.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/512_100_deconv.npy',
                '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/512_500_deconv.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/1024_100_deconv.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Timing Johannes/k53/1024_500_deconv.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_512_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_512_500.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_1024_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_graph/k53_1024_500.npy',
'/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_512_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_512_500.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_1024_100.npy',
 '/media/nel/storage/NEL-LAB Dropbox/NEL/Papers/VolPy_online/CalciumData/fiola_eager/k53_1024_500.npy']
all_files_headers = [fl.split("/")[-1] for fl in all_files]
values = [np.load(fl, allow_pickle=True) for fl in all_files]
#%%
max_len = 3000
dfc = pd.DataFrame()
for i,val in enumerate(all_files_headers):

    filler = [None]* (max_len-len(values[i]))
    dfc[val] = np.concatenate((values[i], filler))
dfs.append(dfc)
#%%
def multiple_dfs(df_list, sheets, file_name, spaces, text):
    try:
        book=openpyxl.load_workbook(file_name)
        print("existing workbook")
    except:
        book=openpyxl.Workbook()
        print("new workbook")
        book.save(file_name)
        
    writer = pd.ExcelWriter(file_name, engine="openpyxl")
    writer.book = book
    writer.sheets = {ws.title: ws for ws in sheets}
    print(writer.sheets)
    
    for i,dataframe in enumerate(df_list):
        row=3
        dataframe.to_excel(writer,sheet_name=sheets[i], startrow=row, startcol=0, index=False, na_rep="NA")
        row = row+ len(dataframe.index)+ spaces+ 1
        # writer.sheets[sheets[i]].append(list(dataframe["times"][0]))
        writer.sheets[sheets[i]].cell(1,1).value=text
        writer.save()
  
#%%
excel_folder = "../../../../media/nel/storage/NEL-LAB Dropbox/NEL/Papers/Nature Methods Resubmission/Data"
text = "Timing data for fiola"
sheets = ["5a", "5b", "5c"]
excel_name = os.path.join(excel_folder, "FIOLA5.xlsx")
multiple_dfs(dfs, sheets, excel_name, 2, "test")

#%% fig 2

